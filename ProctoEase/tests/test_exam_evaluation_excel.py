"""
Tests for Phase E — Recruiter Excel Export.

The endpoint must be a pure RENDERING of the canonical Phase B evaluation:
tests assert ``get_exam_evaluation`` is called exactly ONCE (reuse, not
recalculation) and that generated workbooks round-trip through openpyxl with
scores, persisted risk, the system recommendation, and the Phase D human
recruiter decision preserved — with the decision never altering the
recommendation. All DB access is mocked (no live database).
"""

import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock, MagicMock, patch
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.core.dependencies import get_current_user, get_db
from app.main import app
from app.models.user import User, UserRole
from app.services import exam_evaluation_excel_service


def make_user(role: UserRole = UserRole.RECRUITER) -> User:
    return User(
        id=uuid4(),
        email=f"{role.value}@techcorp.com",
        full_name=f"Test {role.value}",
        role=role,
        tenant_id=uuid4(),
        is_active=True,
    )


def make_evaluation(**overrides):
    """Synthetic evaluation payload shaped exactly like get_exam_evaluation."""
    candidate = {
        "attempt_id": uuid4(),
        "candidate_id": uuid4(),
        "candidate_name": "Sanya Nair",
        "candidate_email": "sanya.nair.04@techcorp.demo",
        "status": "evaluated",
        "started_at": datetime(2026, 8, 27, 12, 0, 0, tzinfo=timezone.utc),
        "submitted_at": datetime(2026, 8, 27, 12, 5, 0, tzinfo=timezone.utc),
        "duration_minutes": 5.0,
        "total_score": 2,
        "max_score": 15,
        "percentage": 13.33,
        "objective_score": 2,
        "objective_max_score": 6,
        "coding_score": 0,
        "coding_max_score": 9,
        "risk_score": 0.8173,
        "risk_level": "critical",
        "risk_available": True,
        "total_violations": 13,
        "high_violations": 2,
        "critical_violations": 1,
        "severe_integrity": True,
        "recommendation": {
            "code": "NOT_RECOMMENDED_BOTH",
            "label": "Not Recommended (Academic & Integrity)",
            "reason": "Score 13.33% is below passing cutoff (50%) with concurrent severe integrity flags (Risk: CRITICAL, score 0.8173).",
        },
        "recruiter_decision": "SHORTLISTED",
        "recruiter_notes": "Candidate selected for further evaluation.",
        "reviewed_by": uuid4(),
        "reviewed_at": datetime(2026, 8, 30, 7, 1, 0, tzinfo=timezone.utc),
        "reviewed_by_email": "recruiter@techcorp.com",
    }
    candidate.update(overrides)
    return {
        "exam_id": uuid4(),
        "exam_title": "Algorithms & Cloud Infrastructure Benchmark (Test 3)",
        "total_attempts": 1,
        "max_score": 15,
        "objective_max_score": 6,
        "coding_max_score": 9,
        "passing_score_pct": 50.0,
        "borderline_max_pct": 60.0,
        "excellence_score_pct": 75.0,
        "candidates": [candidate],
    }


@pytest.fixture
def recruiter():
    return make_user(UserRole.RECRUITER)


def _override(user, mock_db):
    app.dependency_overrides[get_current_user] = lambda: user
    app.dependency_overrides[get_db] = lambda: mock_db


def _workbook_from_bytes(data: bytes):
    return load_workbook(io.BytesIO(data))


class TestExcelWorkbookBuilder:
    def test_valid_xlsx_with_expected_sheets_and_headers(self, recruiter):
        wb = _workbook_from_bytes(
            exam_evaluation_excel_service.build_evaluation_workbook(make_evaluation())[0]
        )
        assert wb.sheetnames == ["Candidate Evaluations", "Export Info"]
        ws = wb["Candidate Evaluations"]
        headers = [c.value for c in ws[1]]
        assert headers[0] == "Candidate Name"
        assert headers[16] == "System Recommendation"
        assert headers[17] == "System Recommendation Reason"
        assert headers[18] == "Recruiter Decision"
        assert headers[20] == "Reviewed By"
        assert headers[21] == "Reviewed At (UTC)"
        # Freeze header row + autofilter on the header range.
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref is not None

    def test_candidate_values_round_trip_with_numeric_scores(self, recruiter):
        wb = _workbook_from_bytes(
            exam_evaluation_excel_service.build_evaluation_workbook(make_evaluation())[0]
        )
        row = [c.value for c in wb["Candidate Evaluations"][2]]
        assert row[0] == "Sanya Nair"
        assert row[1] == "sanya.nair.04@techcorp.demo"
        assert row[2] == "evaluated"
        assert row[3] == 2 and row[4] == 15          # real numbers
        assert row[5] == 13.33                        # percentage numeric
        assert row[6] == 2 and row[8] == 0
        assert row[10] == 0.8173                      # persisted risk, verbatim
        assert row[11] == "critical"
        assert row[12] == 13 and row[13] == 2 and row[14] == 1
        assert row[15] == "Yes"                       # severe integrity

    def test_recommendation_and_decision_preserved_and_separate(self, recruiter):
        wb = _workbook_from_bytes(
            exam_evaluation_excel_service.build_evaluation_workbook(make_evaluation())[0]
        )
        row = [c.value for c in wb["Candidate Evaluations"][2]]
        # System recommendation verbatim from the engine...
        assert row[16] == "NOT_RECOMMENDED_BOTH"
        assert row[17].startswith("Score 13.33% is below passing cutoff")
        # ...while the HUMAN decision is a different, independently persisted value.
        assert row[18] == "SHORTLISTED"
        assert row[19] == "Candidate selected for further evaluation."
        assert row[18] != row[16]  # decision ≠ recommendation, both present

    def test_reviewed_by_email_and_utc_timestamp(self, recruiter):
        wb = _workbook_from_bytes(
            exam_evaluation_excel_service.build_evaluation_workbook(make_evaluation())[0]
        )
        row = [c.value for c in wb["Candidate Evaluations"][2]]
        assert row[20] == "recruiter@techcorp.com"
        assert row[21] == "2026-08-30 07:01:00 UTC"

    def test_null_decision_exports_as_pending(self, recruiter):
        evaluation = make_evaluation(
            recruiter_decision=None, recruiter_notes=None,
            reviewed_by=None, reviewed_at=None, reviewed_by_email=None,
        )
        wb = _workbook_from_bytes(exam_evaluation_excel_service.build_evaluation_workbook(evaluation)[0])
        row = [c.value for c in wb["Candidate Evaluations"][2]]
        assert row[18] == "PENDING"
        assert row[19] is None
        assert row[20] is None

    def test_decision_does_not_alter_recommendation_column(self, recruiter):
        """Same engine recommendation, opposite human decisions — the
        recommendation columns must be byte-identical across exports."""
        base = make_evaluation()
        wb_shortlisted = _workbook_from_bytes(
            exam_evaluation_excel_service.build_evaluation_workbook(base)[0]
        )
        rejected = make_evaluation(recruiter_decision="REJECTED", recruiter_notes="Declined.")
        wb_rejected = _workbook_from_bytes(
            exam_evaluation_excel_service.build_evaluation_workbook(rejected)[0]
        )
        row_a = [c.value for c in wb_shortlisted["Candidate Evaluations"][2]]
        row_b = [c.value for c in wb_rejected["Candidate Evaluations"][2]]
        assert row_a[16] == row_b[16] == "NOT_RECOMMENDED_BOTH"
        assert row_a[17] == row_b[17]
        assert row_a[18] == "SHORTLISTED" and row_b[18] == "REJECTED"

    def test_unicode_survives_round_trip(self, recruiter):
        evaluation = make_evaluation(
            candidate_name="José Müller — Иванов",
            candidate_email="jose.muller@techcorp.demo",
            recruiter_notes="विद्यार्थी का मूल्यांकन «sécurité» – 完成。",
        )
        wb = _workbook_from_bytes(exam_evaluation_excel_service.build_evaluation_workbook(evaluation)[0])
        row = [c.value for c in wb["Candidate Evaluations"][2]]
        assert row[0] == "José Müller — Иванов"
        assert row[19] == "विद्यार्थी का मूल्यांकन «sécurité» – 完成。"

    def test_empty_exam_renders_headers_only_with_info_sheet(self, recruiter):
        evaluation = make_evaluation()
        evaluation["candidates"] = []
        data, row_count = exam_evaluation_excel_service.build_evaluation_workbook(evaluation)
        assert row_count == 0
        wb = _workbook_from_bytes(data)
        ws = wb["Candidate Evaluations"]
        assert ws.max_row == 1  # header row only
        info = wb["Export Info"]
        flat = {info.cell(row=r, column=1).value: info.cell(row=r, column=2).value for r in range(2, info.max_row + 1)}
        assert flat["Total Candidates"] == 0
        assert "not a final hiring decision" in flat["Note"]


class TestReviewedByEmailLookup:
    """Phase E addition to the Phase B payload: reviewer emails must come from
    the SAME single batched user lookup — no N+1 and no second query."""

    def _make_result(self, scalar_one=None, scalars_all=None, rows=None):
        mock = MagicMock()
        mock.scalar_one_or_none = MagicMock(return_value=scalar_one)
        mock.scalars = MagicMock(
            return_value=MagicMock(all=MagicMock(return_value=scalars_all or []))
        )
        mock.all = MagicMock(return_value=rows or [])
        return mock

    @pytest.mark.asyncio
    async def test_reviewer_emails_resolved_in_constant_query_count(self):
        from types import SimpleNamespace

        from app.services import exam_evaluation_service

        exam = SimpleNamespace(id=uuid4(), title="Excel Export Exam")
        reviewer_a, reviewer_b = uuid4(), uuid4()
        attempts = []
        for i, reviewer in enumerate([reviewer_a, reviewer_b, None]):
            attempts.append(
                SimpleNamespace(
                    id=uuid4(),
                    candidate_id=uuid4(),
                    reviewed_by=reviewer,
                    recruiter_decision="SHORTLISTED" if reviewer else None,
                    recruiter_notes="ok" if reviewer else None,
                    reviewed_at=datetime(2026, 8, 30, 7, i, 0, tzinfo=timezone.utc)
                    if reviewer
                    else None,
                    answers={},
                    status="submitted",
                    started_at=datetime(2026, 8, 27, 10, i, 0, tzinfo=timezone.utc),
                    submitted_at=datetime(2026, 8, 27, 11, i, 0, tzinfo=timezone.utc),
                    is_active=True,
                )
            )
        users_rows = [
            (attempts[0].candidate_id, "Cand A", "a@x.demo"),
            (attempts[1].candidate_id, "Cand B", "b@x.demo"),
            (attempts[2].candidate_id, "Cand C", "c@x.demo"),
            (reviewer_a, "Recruiter One", "r1@techcorp.com"),
            (reviewer_b, "Recruiter Two", "r2@techcorp.com"),
        ]
        mock_db = AsyncMock()
        mock_db.execute = AsyncMock(
            side_effect=[
                self._make_result(scalar_one=exam),          # 1. exam
                self._make_result(scalars_all=attempts),      # 2. attempts
                self._make_result(scalars_all=[]),            # 3. questions
                self._make_result(rows=users_rows),           # 4. candidates ∪ reviewers (ONE query)
                self._make_result(rows=[]),                   # 5. grouped events
            ]
        )

        with patch(
            "app.services.exam_evaluation_service.risk_engine.get_exam_risk_scores",
            new_callable=AsyncMock,
        ) as mock_risk:
            mock_risk.return_value = []
            out = await exam_evaluation_service.get_exam_evaluation(
                mock_db, exam.id, uuid4()
            )

        # Query count stays CONSTANT (5) regardless of attempt/reviewer count —
        # the reviewer union reused the existing users query, no N+1.
        assert mock_db.execute.await_count == 5
        emails = [c["reviewed_by_email"] for c in out["candidates"]]
        assert emails == ["r1@techcorp.com", "r2@techcorp.com", None]
        # Recommendation untouched by the Phase E field addition.
        assert out["candidates"][0]["recommendation"]["code"] in {
            "MANUAL_REVIEW", "NOT_RECOMMENDED_ACADEMIC", "NOT_RECOMMENDED_BOTH",
            "INTEGRITY_REVIEW", "SHORTLIST", "STRONG_SHORTLIST",
        }


class TestExcelExportEndpoint:
    def test_recruiter_gets_valid_xlsx_and_calls_evaluation_once(self, recruiter):
        mock_db = AsyncMock()
        evaluation = make_evaluation()
        _override(recruiter, mock_db)
        client = TestClient(app)
        with patch(
            "app.api.v1.reporting.exam_evaluation_service.get_exam_evaluation",
            new_callable=AsyncMock,
        ) as mock_eval:
            mock_eval.return_value = evaluation
            resp = client.get(
                f"/api/v1/exams/{evaluation['exam_id']}/candidate-evaluations/excel"
            )
        try:
            assert resp.status_code == 200
            assert (
                resp.headers["content-type"]
                == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            assert "attachment" in resp.headers.get("content-disposition", "")
            assert resp.headers["content-disposition"].startswith(
                "attachment; filename=\"candidate_evaluations_"
            )
            assert resp.content[:2] == b"PK"  # XLSX is a zip container
            # SINGLE source of truth: the canonical evaluation was reused once.
            mock_eval.assert_awaited_once()
        finally:
            app.dependency_overrides.clear()

    def test_candidate_role_forbidden_403(self):
        candidate_user = make_user(UserRole.CANDIDATE)
        mock_db = AsyncMock()
        _override(candidate_user, mock_db)
        client = TestClient(app)
        try:
            resp = client.get(
                f"/api/v1/exams/{uuid4()}/candidate-evaluations/excel"
            )
            assert resp.status_code == 403
            assert resp.json().get("error_code") == "FORBIDDEN"
        finally:
            app.dependency_overrides.clear()

    def test_cross_tenant_exam_is_404(self, recruiter):
        """get_exam_evaluation raises ExamNotFound for a foreign-tenant exam;
        the endpoint must surface that as 404."""
        mock_db = AsyncMock()
        _override(recruiter, mock_db)
        client = TestClient(app)
        with patch(
            "app.api.v1.reporting.exam_evaluation_service.get_exam_evaluation",
            new_callable=AsyncMock,
        ) as mock_eval:
            from app.core.exceptions import ExamNotFound

            mock_eval.side_effect = ExamNotFound()
            resp = client.get(
                f"/api/v1/exams/{uuid4()}/candidate-evaluations/excel"
            )
        try:
            assert resp.status_code == 404
        finally:
            app.dependency_overrides.clear()

    def test_empty_exam_still_returns_valid_workbook(self, recruiter):
        mock_db = AsyncMock()
        evaluation = make_evaluation()
        evaluation["candidates"] = []
        _override(recruiter, mock_db)
        client = TestClient(app)
        with patch(
            "app.api.v1.reporting.exam_evaluation_service.get_exam_evaluation",
            new_callable=AsyncMock,
        ) as mock_eval:
            mock_eval.return_value = evaluation
            resp = client.get(
                f"/api/v1/exams/{evaluation['exam_id']}/candidate-evaluations/excel"
            )
        try:
            assert resp.status_code == 200
            wb = _workbook_from_bytes(resp.content)
            assert wb.sheetnames == ["Candidate Evaluations", "Export Info"]
            assert wb["Candidate Evaluations"].max_row == 1
        finally:
            app.dependency_overrides.clear()
