"""
Exam-wide candidate evaluation Excel export service (Phase E).

A pure RENDERER over ``exam_evaluation_service.get_exam_evaluation``: the
canonical per-candidate evaluation (scores, persisted risk, violation counts,
deterministic system recommendation, Phase D recruiter decision) is produced
there — batched, tenant-scoped, with risk NEVER recomputed. This module only
shapes that payload into an in-memory XLSX workbook; it contains no business
logic of its own.

Presentation rules:
 - Sheet "Candidate Evaluations": one row per attempt, frozen header row,
   autofilter, numeric scores as real numbers, timestamps as UTC text.
 - Sheet "Export Info": exam identity, generation time (UTC), the engine
   benchmark constants and the row count, for auditability.
 - Recruiter Decision cells are visually emphasised (colored fill per value)
   because they are the final HUMAN judgment; System Recommendation stays
   plain text (code + reason) so the automated suggestion is never presented
   as a decision. NULL decision exports as PENDING.
 - Unicode text (names, emails, notes, reasons) is written natively — XLSX
   is UTF-8 by construction.
"""

from __future__ import annotations

import io
from collections.abc import Mapping
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Presentation constants ───────────────────────────────────────────────────
HEADERS = [
    "Candidate Name",
    "Candidate Email",
    "Attempt Status",
    "Total Score",
    "Max Score",
    "Score Percentage (%)",
    "Objective Score",
    "Objective Max",
    "Coding Score",
    "Coding Max",
    "Risk Score",
    "Risk Level",
    "Total Violations",
    "High Violations",
    "Critical Violations",
    "Severe Integrity",
    "System Recommendation",
    "System Recommendation Reason",
    "Recruiter Decision",
    "Recruiter Notes",
    "Reviewed By",
    "Reviewed At (UTC)",
]

# Column widths per content class (Excel column units).
COL_WIDTHS = [22, 30, 14, 11, 10, 14, 12, 12, 11, 11, 10, 10, 12, 12, 13, 12, 26, 60, 18, 50, 28, 20]
# Columns rendered with wrap_text (long free text).
WRAPPED_COLS = {18, 20}  # System Recommendation Reason, Recruiter Notes (1-based)

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_GREY = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

# Severity-tinted fills (mirroring the recruiter UI pills). Recruiter Decision
# is the colored HUMAN column; System Recommendation stays uncolored.
RISK_FILLS = {
    "critical": PatternFill("solid", fgColor="FECDD3"),
    "high": PatternFill("solid", fgColor="FFEDD5"),
    "medium": PatternFill("solid", fgColor="FEF3C7"),
    "low": PatternFill("solid", fgColor="D1FAE5"),
}
DECISION_FILLS = {
    "SHORTLISTED": PatternFill("solid", fgColor="D1FAE5"),
    "REVIEW": PatternFill("solid", fgColor="FEF3C7"),
    "REJECTED": PatternFill("solid", fgColor="FECDD3"),
    "PENDING": PatternFill("solid", fgColor="F3F4F6"),
}


def _fmt_dt(value: Any) -> str | None:
    """Render an ISO timestamp as an explicit-UTC string, or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is not None:
            value = value.astimezone(timezone.utc)
        return value.strftime("%Y-%m-%d %H:%M:%S UTC")
    return str(value)


def build_evaluation_workbook(evaluation: Mapping[str, Any]) -> tuple[bytes, int]:
    """
    Render one exam's evaluation payload (as returned by
    ``get_exam_evaluation``) into an in-memory XLSX.

    Returns (xlsx_bytes, candidate_row_count).
    """
    wb = Workbook()

    # ── Sheet 1: Candidate Evaluations ───────────────────────────────────
    ws = wb.active
    ws.title = "Candidate Evaluations"
    ws.append(HEADERS)
    for col, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    candidates = evaluation.get("candidates") or []
    for c in candidates:
        row_idx = ws.max_row + 1
        ws.append(
            [
                c.get("candidate_name"),
                c.get("candidate_email"),
                c.get("status"),
                c.get("total_score"),
                c.get("max_score"),
                c.get("percentage"),
                c.get("objective_score"),
                c.get("objective_max_score"),
                c.get("coding_score"),
                c.get("coding_max_score"),
                c.get("risk_score"),
                c.get("risk_level"),
                c.get("total_violations"),
                c.get("high_violations"),
                c.get("critical_violations"),
                "Yes" if c.get("severe_integrity") else "No",
                # SYSTEM RECOMMENDATION — automated decision support, verbatim
                # engine code + reason, deliberately uncolored.
                c.get("recommendation", {}).get("code"),
                c.get("recommendation", {}).get("reason"),
                # RECRUITER DECISION — final HUMAN judgment; NULL → PENDING.
                c.get("recruiter_decision") or "PENDING",
                c.get("recruiter_notes"),
                # Phase E: reviewer email from the batched evaluation lookup.
                c.get("reviewed_by_email"),
                _fmt_dt(c.get("reviewed_at")),
            ]
        )
        # Row styling: wrap long free text; tint risk + decision cells.
        for col in WRAPPED_COLS:
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")
        risk_cell = ws.cell(row=row_idx, column=12)
        risk_cell.fill = RISK_FILLS.get((c.get("risk_level") or "").lower(), PatternFill())
        decision_cell = ws.cell(row=row_idx, column=19)
        decision_cell.fill = DECISION_FILLS.get(
            c.get("recruiter_decision") or "PENDING", PatternFill()
        )
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_idx, column=col).border = BORDER

    # ── Sheet 2: Export Info ─────────────────────────────────────────────
    info = wb.create_sheet("Export Info")
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 60
    generated_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    info.append(["Field", "Value"])
    for cell in info[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row in (
        ("Exam ID", str(evaluation.get("exam_id", ""))),
        ("Exam Title", evaluation.get("exam_title", "")),
        ("Generated At (UTC)", generated_at),
        ("Total Candidates", len(candidates)),
        ("Passing Score (%)", evaluation.get("passing_score_pct")),
        ("Borderline Max (%)", evaluation.get("borderline_max_pct")),
        ("Excellence Score (%)", evaluation.get("excellence_score_pct")),
        (
            "Note",
            "System Recommendation is automated decision support, not a final "
            "hiring decision. Recruiter Decision is the final human judgment "
            "and never alters the System Recommendation.",
        ),
    ):
        info.append(list(row))
    info["A1"].alignment = Alignment(vertical="center")
    info["B9"].alignment = Alignment(wrap_text=True, vertical="top")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue(), len(candidates)
